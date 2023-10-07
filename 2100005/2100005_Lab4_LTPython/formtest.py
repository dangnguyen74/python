from PIL import *
from tkinter import *

from openpyxl import *
wb = load_workbook('Z:\\2111859_Lab4_LTPython\SV.xlsx')
sheet = wb.active
def excel():

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 60
 
    
    sheet.cell(row=1, column=1).value = "MSSV"
    sheet.cell(row=1, column=2).value = "HoTen"
    sheet.cell(row=1, column=3).value = "NgaySinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "SDT"
    sheet.cell(row=1, column=6).value = "HocKy"
    sheet.cell(row=1, column=7).value = "NamHoc"
    sheet.cell(row=1, column=8).value = "MonHoc"

def focus1(event):
    MSSV_field.focus_set()

def focus2(event):
    HoTen_field.focus_set()
 
def focus3(event):
    NgaySinh_field.focus_set()
 
def focus4(event):
    Email_field.focus_set()
 
def focus5(event):
    SDT_field.focus_set()
 
def focus6(event):
    HocKy_field.focus_set()

def focus7(event):
    NamHoc_field.focus_set()

def focus8(event):
    MonHoc_field.focus_set()
 
def clear():
    MSSV_field.delete(0, END)
    HoTen_field.delete(0, END)
    NgaySinh_field.delete(0, END)
    Email_field.delete(0, END)
    SDT_field.delete(0, END)
    HocKy_field.delete(0, END)
    NamHoc_field.delete(0, END)
    MonHoc_field.delete(0, END)

def insert():
     
    # if user not fill any entry
    # then print "empty input"
    if (MSSV_field.get() == "" and
        HoTen_field.get() == "" and
        NgaySinh_field.get() == "" and
        Email_field.get() == "" and
        SDT_field.get() == "" and
        HocKy_field.get() == "" and
        NamHoc_field.get() == ""and
        MonHoc_field.get() == ""): 
             
        print("empty input")
 
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
 
        sheet.cell(row=current_row + 1, column=1).value = MSSV_field.get()
        sheet.cell(row=current_row + 1, column=2).value = HoTen_field.get()
        sheet.cell(row=current_row + 1, column=3).value = NgaySinh_field.get()
        sheet.cell(row=current_row + 1, column=4).value = Email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = SDT_field.get()
        sheet.cell(row=current_row + 1, column=6).value = HocKy_field.get()
        sheet.cell(row=current_row + 1, column=7).value = NamHoc_field.get()
        sheet.cell(row=current_row + 1, column=8).value = MonHoc_field.get()
 
        wb.save('Z:\\2111859_Lab4_LTPython\SV.xlsx')
 
        MSSV_field.focus_set()
        HoTen_field.focus_set()
        NgaySinh_field.focus_set()
        Email_field.focus_set()
        SDT_field.focus_set()
        HocKy_field.focus_set()
        NamHoc_field.focus_set()
        MonHoc_field.focus_set()

        clear()

if __name__ == "__main__":
     
    # create a GUI window
    root = Tk()
 
    # set the background colour of GUI window
    root.configure(background='light green')
 
    # set the title of GUI window
    root.title("Đăng kí hoc phần")
 
    # set the configuration of GUI window
    root.geometry("700x400")
 
    excel()
 
    # create a Form label
    heading = Label(root, text="THÔNG TIN ĐĂNG KÍ HỌC PHẦN", bg="light green")
 
    # create a Name label
    MSSV = Label(root, text="MSSV", bg="light green")
 
 
    # create a Course label
    HoTen = Label(root, text="Họ tên", bg="light green")
 
    # create a Semester label
    NgaySinh = Label(root, text="Ngày sinh", bg="light green")
 
    # create a Form No. label
    Email = Label(root, text="Email", bg="light green")
 
    # create a Contact No. label
    SDT = Label(root, text="SDT", bg="light green")
 
    # create a Email id label
    HocKy = Label(root, text="Học kỳ", bg="light green")
 
    # create a address label
    NamHoc = Label(root, text="Năm học", bg="light green")

    MonHoc = Label(root, text="Môn học", bg="light green")
 
    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    MSSV.grid(row=1, column=0)
    HoTen.grid(row=2, column=0)
    NgaySinh.grid(row=3, column=0)
    Email.grid(row=4, column=0)
    SDT.grid(row=5, column=0)
    HocKy.grid(row=6, column=0)
    NamHoc.grid(row=7, column=0)
    MonHoc.grid(row=8, column=0)
 
    # create a text entry box
    # for typing the information
    MSSV_field = Entry(root)
    HoTen_field = Entry(root)
    NgaySinh_field = Entry(root)
    Email_field = Entry(root)
    SDT_field = Entry(root)
    HocKy_field = Entry(root)
    NamHoc_field = Entry(root)
    MonHoc_field = Entry(root)
 
    # bind method of widget is used for
    # the binding the function with the events
 
    # whenever the enter key is pressed
    # then call the focus1 function
    MSSV_field.bind("<Return>", focus1)
 
    # whenever the enter key is pressed
    # then call the focus2 function
    HoTen_field.bind("<Return>", focus2)
 
    # whenever the enter key is pressed
    # then call the focus3 function
    NgaySinh_field.bind("<Return>", focus3)
 
    # whenever the enter key is pressed
    # then call the focus4 function
    Email_field.bind("<Return>", focus4)
 
    # whenever the enter key is pressed
    # then call the focus5 function
    SDT_field.bind("<Return>", focus5)
 
    # whenever the enter key is pressed
    # then call the focus6 function
    HocKy_field.bind("<Return>", focus6)

    NamHoc_field.bind("<Return>", focus7)

    MonHoc_field.bind("<Return>", focus8)
 
    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    MSSV_field.grid(row=1, column=1, ipadx="100")
    HoTen_field.grid(row=2, column=1, ipadx="100")
    NgaySinh_field.grid(row=3, column=1, ipadx="100")
    Email_field.grid(row=4, column=1, ipadx="100")
    SDT_field.grid(row=5, column=1, ipadx="100")
    HocKy_field.grid(row=6, column=1, ipadx="100")
    NamHoc_field.grid(row=7, column=1, ipadx="100")
    
 
    # call excel function
    excel()
 
    # create a Submit Button and place into the root window
    submit = Button(root, text="Đăng kí", fg="Black", bg="Red", command=insert)
    submit.grid(row=9, column=1)
 
    def kt_mssv():
        while True:
            MSSV = input()
            if MSSV.isdigit() and len(MSSV) == 7:
                return MSSV
            else:
                print("MSSV phải chứa đúng 7 chữ số. Vui lòng nhập lại.")
        
    # start the GUI'''
    root.mainloop()
